import os
import csv
import json
from itertools import groupby, chain

import requests
import us
from openpyxl import load_workbook


data_dir = 'raw-data'
csv_file_out = 'us-states-pop-estimates.csv'

headers = [
    'year',
    'state_fips',
    'state_name',
    'estimate'
]

with open('data-sources.json', 'r') as infile:
    source_data = json.load(infile)


def dl_files():
    ''' Download the files into the `data_dir` if not there already '''

    for years in source_data:

        # grab the URL
        url = source_data[years]

        # use the same filename
        filename = url.split('/')[-1]

        # get a handle to the download path
        filepath = os.path.abspath(
            os.path.join(
                os.path.dirname(__file__),
                data_dir,
                filename
            )
        )

        # skip this one if exists already
        if os.path.exists(filepath):
            continue

        # download xlsx files in binary mode, text in regular write mode
        if url.endswith('xlsx'):
            with requests.get(url, stream=True) as r, open(filepath, 'wb') as outfile:
                outfile.write(r.content)
        else:
            with requests.get(url, stream=True) as r, open(filepath, 'w') as outfile:
                outfile.write(r.text)

        print(f'Downloaded {filename}')

    return data_dir


def get_70_79():
    ''' Parsing the 1970-1979 population estimates text file '''
    
    local_file = os.path.abspath(
        os.path.join(
            os.path.dirname(__file__),
            data_dir,
            source_data['1970-1979'].split('/')[-1]
        )
    )

    # open the file and skip some crufty bits up top
    with open(local_file, 'r') as infile:
        data = infile.read().split("the 1970's.")[-1].splitlines()

    # drill down to just rows of data
    data = [x for x in data if x.strip()][1:]

    # a fixed-width map for each piece of data in each line
    fw_map = {
        'fips': slice(0, 2),
        'state_name': slice(3, 5),
        'age': slice(8, 14),
        'est_1970': slice(15, 22),
        'est_1971': slice(23, 30),
        'est_1972': slice(31, 38),
        'est_1973': slice(39, 46),
        'est_1974': slice(47, 54),
        'est_1975': slice(55, 62),
        'est_1976': slice(63, 70),
        'est_1977': slice(71, 78),
        'est_1978': slice(79, 86),
        'est_1979': slice(87, 94)
    }

    def parse_line(line):
        ''' Turning a line of text into a dict record '''
        return {key: line[fw_map.get(key)].strip() for key in fw_map}

    # need a temporary data holder, as this data is broken out by age
    # and will need to group later
    data_tmp = []

    # parse each line into a new record in the list
    for line in data:
        data_tmp.append(parse_line(line))

    # the list to hold the finished data
    data_out = []

    # group the data by fips code
    for fips, group in groupby(data_tmp, lambda x: x.get('fips')):
        
        # grab a list of the rows in this group
        state_rows = list(group)

        # and use the first item to get a list of keys tied
        # to population estimate values
        keys_of_interest = [x for x in state_rows[0].keys() if 'est' in x]

        # need zero padding to look up in `us`
        fips = fips.zfill(2)

        # grab state data based on fips
        state_data = us.states.lookup(fips)

        # hey look it's the state's full name
        state_name = state_data.name

        # gonna keep track of state totals in a dict
        state_totals = {}

        # iterate through each line of data
        for age_group in state_rows:

            # loop over the keys for pop estimates
            for key in keys_of_interest:

                # create a counter if not exists already
                if not state_totals.get(key):
                    state_totals[key] = 0

                # increment the counter for this year
                state_totals[key] += int(age_group[key])

        # finally, loop over the object just created
        for yearkey in state_totals:

            # grab a handle to the actual year
            year = int(yearkey.split('_')[-1])

            # and the total for this year
            total = state_totals[yearkey]

            # and add this record to the main data list
            data_out.append(
                dict(
                    zip(
                        headers,
                        [year, fips, state_name, total]
                    )
                )
            )

    # sort by fips, then year, and return
    data_out.sort(
        key=lambda x: (
            x.get('state_fips'),
            x.get('year')
        )
    )

    return data_out


def get_80_89():
    ''' Parsing the 1980-1989 population estimates text file '''

    local_file = os.path.abspath(
        os.path.join(
            os.path.dirname(__file__),
            data_dir,
            source_data['1980-1989'].split('/')[-1]
        )
    )

    # open the file and peel off some cruft at the top
    with open(local_file, 'r') as infile:
        data = infile.read().split('issued 11/93')

    # this text file is "paginated" into two parts 🤌
    # so need to break into top part and bottom part
    pt1 = [x for x in data[1].split('Intercensal Estimates')[0].splitlines() if x.strip()][1:]

    pt2 = [x for x in data[2].splitlines() if x.strip()][1:]

    def parse_line(line):
        ''' Given a line, from either part, parse it into a generic list '''

        # grab state abbreviation, to start with
        state_abbr = line[:2].strip()

        # bail if it's the US record
        if state_abbr == 'US':
            return []

        # look up FIPS and state full name with `us`
        state = us.states.lookup(state_abbr)
        fips = state.fips
        state_name = state.name

        # pull out the values for the five years of data
        date1 = line[4:13].strip()
        date2 = line[14:23].strip()
        date3 = line[24:33].strip()
        date4 = line[34:43].strip()
        date5 = line[44:53].strip()

        return [
            state_name,
            fips,
            date1,
            date2,
            date3,
            date4,
            date5
        ]

    def parse_chunk(part, range_):
        ''' given a part of this file and the range of years, parse
            each line into its atomic record, append to a list and finally return the list of records
         '''
        data_out = []
        for line in part:
            parsed = [x for x in parse_line(line) if x]
            if parsed:
                state = parsed[0]
                fips = parsed[1]

                values = list(zip(range_, parsed[2:]))
                for pair in values:
                    year, est = pair
                    data_out.append(
                        dict(
                            zip(
                                headers,
                                [year, fips, state, est]
                            )
                        )
                    )

        return data_out

    # chunk 1 covers 1980-85
    pt1_data = parse_chunk(pt1, range(1980, 1985))

    # chunk 2 covers 1985-1990 (tho we're leaving 1990 off)
    pt2_data = parse_chunk(pt2, range(1985, 1991))

    data_out = pt1_data + pt2_data

    # sort by fips, then year, and return
    data_out.sort(
        key=lambda x: (
            x.get('state_fips'),
            x.get('year')
        )
    )

    return data_out


def get_90_99():
    ''' Parsing the 1990-1999 population estimates text file '''

    local_file = os.path.abspath(
        os.path.join(
            os.path.dirname(__file__),
            data_dir,
            source_data['1990-1999'].split('/')[-1]
        )
    )

    # open the file and do a little surgery to isolate the first data table
    with open(local_file, 'r') as infile:
        data = infile.read().split('-----Population-----')[1].split('-----Period Births-----')[0].splitlines()[1:]

    # a fixed-width map for each line
    fw_map = {
        'block_num': slice(0, 6),
        'fips': slice(7, 12),
        'state': slice(13, 37),
        'est_1999': slice(38, 47),
        'est_1998': slice(48, 57),
        'est_1997': slice(58, 67),
        'est_1996': slice(68, 77),
        'est_1995': slice(78, 87),
        'est_1994': slice(88, 97),
        'est_1993': slice(98, 107),
        'est_1992': slice(108, 117),
        'est_1991': slice(118, 127),
        'est_1990': slice(128, 137)
    }

    # a list to keep track of the parsed data
    data_out = []

    # loop over the data
    for row in data:

        # skip blank/decorative rows
        if not row.strip() or row.strip().startswith('---'):
            continue

        # grab block and fips
        block = row[fw_map.get('block_num')].strip()
        fips = row[fw_map.get('fips')].strip()

        # skip US
        if fips == '00':
            continue

        # grab state value
        state = row[fw_map.get('state')].strip()

        # skip if no state or crufty row
        if not state or state.startswith('---') or state.startswith('Area '):
            continue

        # loop over the range of years
        for year in range(1990, 2000):
            # grab the total for this year
            total = row[fw_map.get(f'est_{year}')].strip()

            # and add to this list
            data_out.append(
                dict(
                    zip(
                        headers,
                        [year, fips, state, total]
                    )
                )
            )

    # sort by fips, then year, and return
    data_out.sort(
        key=lambda x: (
            x.get('state_fips'),
            x.get('year')
        )
    )

    return data_out


def get_00_09():
    ''' Parsing the 2000-2009 population estimates CSV file '''

    local_file = os.path.abspath(
        os.path.join(
            os.path.dirname(__file__),
            data_dir,
            source_data['2000-2009'].split('/')[-1]
        )
    )

    # list to keep track of parsed data
    data_out = []

    # open the CSV
    with open(local_file, 'r') as infile:

        # load it into a list reader
        reader = list(csv.reader(infile))

        # filter to get just rows for states (start w/ a period)
        data = [x for x in reader if x[0].startswith('.')]

        # loop over the list of rows
        for row in data:

            # grab the state name
            state = row[0].lstrip('.').strip()

            # and look up the FIPS code with `us`
            state_data = us.states.lookup(state)
            fips = state_data.fips

            # pair up years with the associated pop estimate in the row
            # after killing out commas in totals and parsing as int
            estimates = list(zip(range(2000, 2010), [int(x.replace(',', '')) for x in row[2:]]))

            # for each pair, isolate the year and estimate and bundle
            # as a record for the tracking list
            for pair in estimates:
                year, est = pair
                data_out.append(
                    dict(
                        zip(
                            headers,
                            [year, fips, state, est]
                        )
                    )
                )

    # sort by fips, then year, and return
    data_out.sort(
        key=lambda x: (
            x.get('state_fips'),
            x.get('year')
        )
    )

    return data_out


def get_10_19():
    ''' Parsing the 2010-2019 population estimates xlsx file '''

    local_file = os.path.abspath(
        os.path.join(
            os.path.dirname(__file__),
            data_dir,
            source_data['2010-2020'].split('/')[-1]
        )
    )

    # list to keep track of parsed data
    data_out = []

    # load up the workbook and the correct worksheet
    wb = load_workbook(filename=local_file)
    sheet = wb['NST01']

    # filter to just get rows with states (periods prepended)
    data = [x for x in sheet.values if x[0] and x[0].startswith('.')]

    # loop over the state data
    for row in data:

        # get the state name
        state = row[0].lstrip('.').strip()

        # sorry PR 🙁
        if state.lower().startswith('puerto'):
            continue

        # use state name to look up FIPS using `us`
        state_data = us.states.lookup(state)
        fips = state_data.fips

        # zip the range with the remainder of the row
        for pair in zip(range(2010, 2020), row[3:]):

            # isolate year and estimate and package up as a record
            # for the tracking list
            year, est = pair

            data_out.append(
                dict(
                    zip(
                        headers,
                        [year, fips, state, est]
                    )
                )
            )

    # sort by fips, then year, and return
    data_out.sort(
        key=lambda x: (
            x.get('state_fips'),
            x.get('year')
        )
    )

    return data_out



def get_20_22():
    ''' Parsing the 2020-2022 population estimates xlsx file '''
    local_file = os.path.abspath(
        os.path.join(
            os.path.dirname(__file__),
            data_dir,
            source_data['2020-2022'].split('/')[-1]
        )
    )

    # list to keep track of parsed data
    data_out = []

    # load up the workbook and the correct worksheet
    wb = load_workbook(filename=local_file)
    sheet = wb['NST-EST2022-POP']

    # filter to just get rows with states (periods prepended)
    data = [x for x in sheet.values if x[0] and x[0].startswith('.')]

    # loop over the data
    for row in data:

        # grab state name
        state = row[0].lstrip('.').strip()

        # sorry PR 🙁
        if state.lower().startswith('puerto'):
            continue

        # use state name to look up FIPS with `us`
        state_data = us.states.lookup(state)
        fips = state_data.fips

        # pair up the years with the values in the rest of the row
        for pair in zip(range(2020, 2023), row[2:]):

            # grab year and pop estimate
            year, est = pair

            # and send this record to the tracking list
            data_out.append(
                dict(
                    zip(
                        headers,
                        [year, fips, state, est]
                    )
                )
            )

    # sort by fips, then year, and return
    data_out.sort(
        key=lambda x: (
            x.get('state_fips'),
            x.get('year')
        )
    )

    return data_out


def gather_data():
    ''' Grabbing data from each population file '''

    # itertools ftw! just slamming these lists together
    data = list(chain(
        get_70_79(),
        get_80_89(),
        get_90_99(),
        get_00_09(),
        get_10_19(),
        get_20_22()
    ))

    # sort by fips, then year, and return
    data.sort(
        key=lambda x: (
            x.get('state_fips'),
            x.get('year')
        )
    )

    return data



if __name__ == '__main__':

    # download any missing files
    dl_files()

    # gather up the data into a single list
    data = gather_data()

    # and write to file
    with open(csv_file_out, 'w') as outfile:
        writer = csv.DictWriter(outfile, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)

        print(f'Wrote {csv_file_out}')
